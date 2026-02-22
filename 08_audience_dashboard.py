# scripts/08_audience_dashboard.py
"""
===============================================================================
 AUDIENCE PERFORMANCE DASHBOARD – REAL INSIGHTS FROM TARGET AUDIENCES
===============================================================================
Adds to your Marketing Master Dashboard:
   • Audience Summary – revenue, ROI, CTR, CVR per audience (with Indian formatting)
   • Audience Trend – monthly revenue by audience (heatmap ready)
   • Audience Channel Matrix – best channel per audience
   • Top Audiences – ranked with Indian formatting
===============================================================================
"""

import pandas as pd
import numpy as np
import os
import sqlite3
import traceback
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule

# ------------------------------------------------------------
# HELPER FUNCTION – FORMAT INDIAN NUMBERS
# ------------------------------------------------------------
def format_indian(num):
    """Convert large numbers to Indian readable format (K, L, Cr)"""
    if pd.isna(num) or num == 0:
        return "0"
    if num >= 10**7:  # 1 Crore
        return f"{num/10**7:.1f}Cr"
    elif num >= 10**5:  # 1 Lakh
        return f"{num/10**5:.1f}L"
    elif num >= 10**3:  # 1 Thousand
        return f"{num/10**3:.1f}K"
    else:
        return str(int(num))

try:
    # ------------------------------------------------------------
    # CONFIGURATION
    # ------------------------------------------------------------
    INPUT_PATH = os.path.join('..', 'output', 'Marketing_Master_Dashboard.xlsx')
    OUTPUT_PATH = os.path.join('..', 'output', 'Marketing_Master_Dashboard.xlsx')

    print("=" * 60)
    print("📊 AUDIENCE PERFORMANCE DASHBOARD")
    print("=" * 60)

    # ------------------------------------------------------------
    # LOAD CLEANED DATA
    # ------------------------------------------------------------
    db_path = os.path.join('..', 'output', 'marketing.db')
    if os.path.exists(db_path):
        conn = sqlite3.connect(db_path)
        df = pd.read_sql_query("SELECT * FROM campaigns", conn)
        conn.close()
        print("✅ Loaded cleaned data from SQLite.")
    else:
        excel_path = os.path.join('..', 'output', 'cleaned_marketing_data.xlsx')
        if not os.path.exists(excel_path):
            raise FileNotFoundError("No cleaned data found. Run 01_load_clean.py first.")
        df = pd.read_excel(excel_path)
        print("✅ Loaded cleaned data from Excel.")

    df['Date'] = pd.to_datetime(df['Date'])
    print(f"📊 Total rows: {len(df):,}")

    # ------------------------------------------------------------
    # AUDIENCE SUMMARY
    # ------------------------------------------------------------
    print("\n🔍 Computing audience summary...")
    audience_summary = df.groupby('Target_Audience').agg({
        'Revenue': 'sum',
        'Spend': 'sum',
        'Conversion': 'sum',
        'Clicks': 'sum',
        'Impressions': 'sum',
        'ROI': 'mean',
        'CTR': 'mean',
        'CVR': 'mean'
    }).reset_index()

    audience_summary['ROI'] = audience_summary['ROI'].round(1)
    audience_summary['CTR'] = audience_summary['CTR'].round(2)
    audience_summary['CVR'] = audience_summary['CVR'].round(2)
    # Format Revenue and Spend using Indian units
    audience_summary['Revenue_fmt'] = audience_summary['Revenue'].apply(format_indian)
    audience_summary['Spend_fmt'] = audience_summary['Spend'].apply(format_indian)
    # Drop raw numeric columns if you don't need them, but keep for sorting
    # We'll keep both for now; later we'll write only the formatted ones.

    audience_summary = audience_summary.sort_values('Revenue', ascending=False)

    # ------------------------------------------------------------
    # AUDIENCE TREND (Monthly Revenue)
    # ------------------------------------------------------------
    print("🔍 Computing audience trends...")
    df['Month'] = df['Date'].dt.to_period('M')
    audience_trend = df.groupby(['Target_Audience', 'Month'])['Revenue'].sum().reset_index()
    audience_trend['Month'] = audience_trend['Month'].astype(str)
    audience_trend_pivot = audience_trend.pivot(index='Target_Audience', columns='Month', values='Revenue').fillna(0).round(0).astype(int)

    # ------------------------------------------------------------
    # AUDIENCE CHANNEL MATRIX (Best channel per audience by revenue)
    # ------------------------------------------------------------
    print("🔍 Computing audience‑channel matrix...")
    audience_channel = df.groupby(['Target_Audience', 'Channel'])['Revenue'].sum().reset_index()
    top_channel_per_audience = audience_channel.loc[audience_channel.groupby('Target_Audience')['Revenue'].idxmax()].reset_index(drop=True)
    top_channel_per_audience = top_channel_per_audience[['Target_Audience', 'Channel', 'Revenue']]
    top_channel_per_audience['Revenue_fmt'] = top_channel_per_audience['Revenue'].apply(format_indian)

    # ------------------------------------------------------------
    # LOAD EXISTING DASHBOARD AND ADD NEW SHEETS
    # ------------------------------------------------------------
    print("\n💾 Adding audience sheets to dashboard...")

    if not os.path.exists(INPUT_PATH):
        raise FileNotFoundError(f"Dashboard file not found: {INPUT_PATH}. Run 07_dashboard.py first.")

    wb = load_workbook(INPUT_PATH)

    # Remove old RFM/cohort sheets if they exist
    for sheet in ['Customer RFM', 'Segment Summary', 'Cohort Retention', 'Cohort Data']:
        if sheet in wb.sheetnames:
            wb.remove(wb[sheet])
            print(f"🗑️ Removed old sheet: {sheet}")

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    currency_format = '₹#,##0'
    percent_format = '0.00%'

    def format_sheet(ws):
        # Headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        # Auto‑fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        # Borders for data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    # --- Sheet 1: Audience Summary (with formatted numbers) ---
    summary_display = audience_summary[['Target_Audience', 'Revenue_fmt', 'Spend_fmt', 'Conversion', 'ROI', 'CTR', 'CVR']].copy()
    summary_display.columns = ['Target_Audience', 'Revenue', 'Spend', 'Conversion', 'ROI', 'CTR', 'CVR']
    ws_sum = wb.create_sheet('Audience Summary')
    for r_idx, row in enumerate(dataframe_to_rows(summary_display, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_sum.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1:
                col_name = summary_display.columns[c_idx-1]
                if col_name in ['ROI', 'CTR', 'CVR']:
                    cell.number_format = percent_format
    format_sheet(ws_sum)

    # --- Sheet 2: Audience Trend (heatmap ready) ---
    ws_trend = wb.create_sheet('Audience Trend')
    for r_idx, row in enumerate(dataframe_to_rows(audience_trend_pivot.reset_index(), index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_trend.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1 and c_idx > 1:
                cell.number_format = currency_format
    format_sheet(ws_trend)

    # Add color scale to trend cells
    last_col = ws_trend.max_column
    last_row = ws_trend.max_row
    if last_col > 1 and last_row > 1:
        color_scale = ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
        ws_trend.conditional_formatting.add(f'B2:{chr(64+last_col)}{last_row}', color_scale)

    # --- Sheet 3: Audience Channel Matrix (with formatted revenue) ---
    channel_display = top_channel_per_audience[['Target_Audience', 'Channel', 'Revenue_fmt']].copy()
    channel_display.columns = ['Target_Audience', 'Channel', 'Revenue']
    ws_matrix = wb.create_sheet('Audience Channel')
    for r_idx, row in enumerate(dataframe_to_rows(channel_display, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_matrix.cell(row=r_idx, column=c_idx, value=value)
    format_sheet(ws_matrix)

    # --- Sheet 4: Top Audiences (ranked, with formatted revenue) ---
    top_aud = audience_summary.head(10)[['Target_Audience', 'Revenue_fmt', 'ROI', 'CTR', 'CVR']].copy()
    if len(top_aud) > 0:
        top_aud['Rank'] = range(1, len(top_aud) + 1)
        cols = ['Rank'] + [col for col in top_aud.columns if col != 'Rank']
        top_aud = top_aud[cols]
        # Rename Revenue_fmt to Revenue for display
        top_aud = top_aud.rename(columns={'Revenue_fmt': 'Revenue'})
    else:
        top_aud = pd.DataFrame({'Rank': [], 'Target_Audience': [], 'Revenue': [], 'ROI': [], 'CTR': [], 'CVR': []})
    ws_top = wb.create_sheet('Top Audiences')
    for r_idx, row in enumerate(dataframe_to_rows(top_aud, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_top.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1 and c_idx in [4,5]:
                cell.number_format = percent_format
    format_sheet(ws_top)

    wb.save(OUTPUT_PATH)
    print(f"✅ Audience sheets added to:\n   {OUTPUT_PATH}")

    # ------------------------------------------------------------
    # PRINT KEY INSIGHTS
    # ------------------------------------------------------------
    print("\n" + "=" * 60)
    print("🔑 KEY AUDIENCE INSIGHTS")
    print("=" * 60)

    if len(audience_summary) > 0:
        top_audience = audience_summary.iloc[0]['Target_Audience']
        top_rev = audience_summary.iloc[0]['Revenue']
        top_roi = audience_summary.iloc[0]['ROI']
        print(f"\n🏆 Top Audience by Revenue: {top_audience} – {format_indian(top_rev)} (ROI {top_roi:.1f}%)")

        best_roi_aud = audience_summary.loc[audience_summary['ROI'].idxmax()]
        print(f"📈 Highest ROI Audience: {best_roi_aud['Target_Audience']} – {best_roi_aud['ROI']:.1f}%")
    else:
        print("\n⚠️ No audience data found.")

    print("\n✅ Audience dashboard complete!")

except Exception as e:
    print("\n❌ AN ERROR OCCURRED:")
    traceback.print_exc()

finally:
    input("\nPress Enter to close this window...")