# scripts/10_propensity_scoring.py
"""
===============================================================================
 LAYER 4: PROPENSITY SCORING – PREDICT CAMPAIGN SUCCESS
===============================================================================
Adds to your Marketing Master Dashboard:
   • Propensity Score – predicted probability of high ROI for each campaign
   • Feature Importance – what factors drive success
   • Model Performance – accuracy, precision, recall
===============================================================================
"""

import pandas as pd
import numpy as np
import os
import sqlite3
import traceback
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix
from sklearn.preprocessing import LabelEncoder
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# HELPER FUNCTION – FORMAT INDIAN NUMBERS
# ------------------------------------------------------------
def format_indian(num):
    if pd.isna(num) or num == 0:
        return "0"
    if num >= 10**7:
        return f"{num/10**7:.1f}Cr"
    elif num >= 10**5:
        return f"{num/10**5:.1f}L"
    elif num >= 10**3:
        return f"{num/10**3:.1f}K"
    else:
        return str(int(num))

try:
    print("=" * 60)
    print("📊 LAYER 4: PROPENSITY SCORING")
    print("=" * 60)

    # ------------------------------------------------------------
    # LOAD CLEANED DATA
    # ------------------------------------------------------------
    db_path = os.path.join('..', 'output', 'marketing.db')
    if os.path.exists(db_path):
        conn = sqlite3.connect(db_path)
        df = pd.read_sql_query("SELECT * FROM campaigns", conn)
        conn.close()
        print("✅ Loaded cleaned data from SQLite.")
    else:
        excel_path = os.path.join('..', 'output', 'cleaned_marketing_data.xlsx')
        df = pd.read_excel(excel_path)
        print("✅ Loaded cleaned data from Excel.")

    print(f"📊 Total rows: {len(df):,}")

    # ------------------------------------------------------------
    # PREPARE FEATURES AND TARGET
    # ------------------------------------------------------------
    # Define a binary target: high ROI (1) if ROI > median ROI, else 0
    median_roi = df['ROI'].median()
    df['High_ROI'] = (df['ROI'] > median_roi).astype(int)

    # Features: we'll use categorical and numeric columns
    feature_cols = ['Channel', 'Target_Audience', 'Creative_Type', 'Spend', 'Impressions', 'Clicks', 'CTR']
    # Ensure all exist
    feature_cols = [col for col in feature_cols if col in df.columns]

    # Handle categorical variables with label encoding
    le_dict = {}
    X = df[feature_cols].copy()
    for col in ['Channel', 'Target_Audience', 'Creative_Type']:
        if col in X.columns:
            le = LabelEncoder()
            X[col] = le.fit_transform(X[col].astype(str))
            le_dict[col] = le

    # Drop rows with missing values (should be none)
    X = X.dropna()
    y = df.loc[X.index, 'High_ROI']

    print(f"✅ Features prepared: {feature_cols}")
    print(f"📊 Training samples: {len(X)}")

    # ------------------------------------------------------------
    # TRAIN/TEST SPLIT
    # ------------------------------------------------------------
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42, stratify=y)

    # ------------------------------------------------------------
    # TRAIN RANDOM FOREST CLASSIFIER
    # ------------------------------------------------------------
    rf = RandomForestClassifier(n_estimators=100, random_state=42, class_weight='balanced')
    rf.fit(X_train, y_train)

    # ------------------------------------------------------------
    # EVALUATE
    # ------------------------------------------------------------
    y_pred = rf.predict(X_test)
    accuracy = accuracy_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred)
    recall = recall_score(y_test, y_pred)
    f1 = f1_score(y_test, y_pred)

    # Feature importance
    importances = pd.DataFrame({
        'Feature': feature_cols,
        'Importance': rf.feature_importances_
    }).sort_values('Importance', ascending=False)

    print("\n📈 Model Performance:")
    print(f"   Accuracy:  {accuracy:.2%}")
    print(f"   Precision: {precision:.2%}")
    print(f"   Recall:    {recall:.2%}")
    print(f"   F1 Score:  {f1:.2%}")

    # ------------------------------------------------------------
    # PREDICT PROPENSITY FOR ALL CAMPAIGNS
    # ------------------------------------------------------------
    # Re‑encode using the same encoders
    X_all = df[feature_cols].copy()
    for col, le in le_dict.items():
        X_all[col] = le.transform(X_all[col].astype(str))

    # Predict probabilities (chance of high ROI)
    probs = rf.predict_proba(X_all)[:, 1]  # probability of class 1 (high ROI)

    # Add to dataframe
    df_result = df[['Campaign_Id', 'Campaign_Name', 'Channel', 'ROI']].copy()
    df_result['Propensity_Score'] = probs
    df_result['Predicted_High_ROI'] = (probs >= 0.5).astype(int)
    df_result['Propensity_Score'] = df_result['Propensity_Score'].round(3)

    # Sort by propensity (highest first)
    df_result = df_result.sort_values('Propensity_Score', ascending=False).reset_index(drop=True)
    df_result['Rank'] = range(1, len(df_result) + 1)

    # ------------------------------------------------------------
    # GENERATE INSIGHTS
    # ------------------------------------------------------------
    insights_lines = [
        "🔮 PROPENSITY INSIGHTS",
        "─────────────────────────────────────",
        f"• Model accuracy: {accuracy:.1%} – predicts high‑ROI campaigns with {accuracy:.1%} reliability.",
        f"• Top features driving success:",
    ]
    for i, row in importances.head(3).iterrows():
        insights_lines.append(f"   - {row['Feature']}: {row['Importance']:.1%}")
    insights_lines.append(f"• Campaigns with >70% propensity: {len(df_result[df_result['Propensity_Score']>0.7])} campaigns identified.")
    insights_lines.append("• Use this sheet to prioritize future campaigns with highest success probability.")
    insights_lines.append("")

    # ------------------------------------------------------------
    # LOAD DASHBOARD AND ADD NEW SHEET
    # ------------------------------------------------------------
    INPUT_PATH = os.path.join('..', 'output', 'Marketing_Master_Dashboard.xlsx')
    OUTPUT_PATH = INPUT_PATH

    if not os.path.exists(INPUT_PATH):
        raise FileNotFoundError(f"Dashboard file not found: {INPUT_PATH}. Run 07_dashboard.py first.")

    wb = load_workbook(INPUT_PATH)

    # Remove old propensity sheet if exists
    if 'Propensity Score' in wb.sheetnames:
        wb.remove(wb['Propensity Score'])
        print("🗑️ Removed old Propensity Score sheet.")

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    currency_format = '₹#,##0'
    percent_format = '0.00%'

    def format_sheet(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    # Create new sheet
    ws_prop = wb.create_sheet('Propensity Score')

    # Write insights at top
    for i, line in enumerate(insights_lines, start=1):
        cell = ws_prop.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14, color='1F4E79')
        elif line.startswith('─'):
            cell.font = Font(size=10)
        else:
            cell.font = Font(size=11)
        ws_prop.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    # Write feature importance table (right side)
    start_row_imp = 1
    ws_prop.cell(row=start_row_imp, column=7, value="📈 FEATURE IMPORTANCE").font = Font(bold=True, size=12, color='1F4E79')
    ws_prop.merge_cells(start_row=start_row_imp, start_column=7, end_row=start_row_imp, end_column=8)
    ws_prop.cell(row=start_row_imp+1, column=7, value="Feature").font = Font(bold=True)
    ws_prop.cell(row=start_row_imp+1, column=8, value="Importance").font = Font(bold=True)
    for i, row in importances.iterrows():
        ws_prop.cell(row=start_row_imp+2+i, column=7, value=row['Feature'])
        cell = ws_prop.cell(row=start_row_imp+2+i, column=8, value=row['Importance'])
        cell.number_format = percent_format

    # Write propensity table starting after insights
    start_row_data = len(insights_lines) + 2
    cols_display = ['Rank', 'Campaign_Id', 'Campaign_Name', 'Channel', 'ROI', 'Propensity_Score', 'Predicted_High_ROI']
    disp_df = df_result[cols_display].copy()
    disp_df['ROI'] = disp_df['ROI'].round(1)
    disp_df.columns = ['Rank', 'Campaign ID', 'Campaign Name', 'Channel', 'ROI %', 'Propensity Score', 'Predicted High ROI']

    for r_idx, row in enumerate(dataframe_to_rows(disp_df, index=False, header=True), start=start_row_data):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_prop.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row_data:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                if c_idx == 5:  # ROI %
                    cell.number_format = percent_format
                elif c_idx == 6:  # Propensity Score
                    cell.number_format = percent_format
                cell.border = thin_border

    format_sheet(ws_prop)

    wb.save(OUTPUT_PATH)
    print(f"✅ Propensity Score sheet added to:\n   {OUTPUT_PATH}")

except Exception as e:
    print("\n❌ AN ERROR OCCURRED:")
    traceback.print_exc()

finally:
    input("\nPress Enter to close this window...")