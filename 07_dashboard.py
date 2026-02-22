# scripts/07_dashboard.py
"""
===============================================================================
 MASTER MARKETING DASHBOARD – PROFESSIONAL WITH INSIGHTS & CHARTS
===============================================================================
Combines all analysis into a single Excel file with:
   • Executive Summary (KPI cards)
   • Key Insights (narrative takeaways from all 10 questions)
   • All analysis sheets with formatted data and embedded bar charts
   • Conditional formatting, number formats, and clean layout
===============================================================================
"""

import pandas as pd
import os
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

# -----------------------------------------------------------------------------
# HELPER FUNCTIONS
# -----------------------------------------------------------------------------
def add_chart(ws, title, data_col, category_col, chart_position):
    """Add a bar chart to the worksheet."""
    chart = BarChart()
    chart.title = title
    chart.style = 10
    # Data range (values)
    data = Reference(ws, min_col=data_col, min_row=2, max_row=ws.max_row, max_col=data_col)
    # Categories range
    cats = Reference(ws, min_col=category_col, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.height = 7.5
    chart.width = 12
    ws.add_chart(chart, chart_position)

# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------
try:
    print("=" * 60)
    print("📊 BUILDING PROFESSIONAL MASTER DASHBOARD")
    print("=" * 60)

    output_dir = os.path.join('..', 'output')
    dashboard_path = os.path.join(output_dir, 'Marketing_Master_Dashboard.xlsx')

    # ------------------------------------------------------------
    # 1. LOAD ALL EXISTING OUTPUTS
    # ------------------------------------------------------------
    files = {
        'Campaign Analysis': 'campaign_analysis.xlsx',
        'ML Insights': 'ml_insights.xlsx',
        'Audience Segments': 'audience_segments.xlsx',
        'Budget Optimization': 'budget_optimization.xlsx',
        'Forecast': 'forecast.xlsx'
    }

    data = {}
    for name, fname in files.items():
        path = os.path.join(output_dir, fname)
        if os.path.exists(path):
            try:
                xl = pd.ExcelFile(path)
                sheets = {}
                for sheet in xl.sheet_names:
                    sheets[sheet] = pd.read_excel(path, sheet_name=sheet)
                data[name] = sheets
                print(f"✅ Loaded {name} ({len(sheets)} sheets)")
            except Exception as e:
                print(f"⚠️ Could not load {name}: {e}")
        else:
            print(f"⚠️ {name} not found – skipping")

    if not data:
        raise Exception("No data files found. Run previous scripts first.")

    # ------------------------------------------------------------
    # 2. CREATE NEW WORKBOOK
    # ------------------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    currency_format = '₹#,##0'
    percent_format = '0.00%'

    # ------------------------------------------------------------
    # 3. EXECUTIVE SUMMARY SHEET (KPI CARDS)
    # ------------------------------------------------------------
    ws_sum = wb.create_sheet('Executive Summary', 0)

    # Gather KPIs from various sources
    kpis = []

    if 'Campaign Analysis' in data and 'Channel Performance' in data['Campaign Analysis']:
        df_chan = data['Campaign Analysis']['Channel Performance']
        total_rev = df_chan['Revenue'].sum()
        total_spend = df_chan['Spend'].sum()
        overall_roi = (total_rev - total_spend) / total_spend * 100
        kpis.extend([
            ('Total Revenue', f'₹{total_rev:,.0f}'),
            ('Total Spend', f'₹{total_spend:,.0f}'),
            ('Overall ROI', f'{overall_roi:.1f}%')
        ])

    if 'Campaign Analysis' in data and 'Top Revenue' in data['Campaign Analysis']:
        top_camp = data['Campaign Analysis']['Top Revenue'].iloc[0]
        kpis.append(('Top Campaign', top_camp['Campaign_Name']))
        kpis.append(('Top Campaign Revenue', f'₹{top_camp["Revenue"]:,.0f}'))
        kpis.append(('Top Campaign ROI', f'{top_camp["ROI"]:.1f}%'))

    if 'Campaign Analysis' in data and 'Channel Performance' in data['Campaign Analysis']:
        best_chan = data['Campaign Analysis']['Channel Performance'].iloc[0]
        kpis.append(('Best Channel', best_chan['Channel']))
        kpis.append(('Best Channel ROI', f'{best_chan["ROI"]:.1f}%'))

    if 'ML Insights' in data and 'Cluster Profiles' in data['ML Insights']:
        df_clust = data['ML Insights']['Cluster Profiles']
        high_roi = df_clust[df_clust['Label'] == 'High ROI']
        if not high_roi.empty:
            kpis.append(('High ROI Cluster Size', high_roi['Size'].iloc[0]))

    if 'Forecast' in data and 'Forecast' in data['Forecast']:
        df_fc = data['Forecast']['Forecast']
        future_total = df_fc['yhat'].sum()
        kpis.append(('Forecast Next 90 Days', f'₹{future_total:,.0f}'))

    # Write KPIs to sheet (two columns)
    for i, (metric, value) in enumerate(kpis, start=1):
        ws_sum.cell(row=i, column=1, value=metric).font = Font(bold=True)
        ws_sum.cell(row=i, column=2, value=value).alignment = Alignment(horizontal='right')
        ws_sum.cell(row=i, column=1).border = thin_border
        ws_sum.cell(row=i, column=2).border = thin_border

    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 25

    # ------------------------------------------------------------
    # 4. KEY INSIGHTS SHEET – ALL 10 QUESTIONS ANSWERED
    # ------------------------------------------------------------
    ws_insights = wb.create_sheet('Key Insights', 1)

    insights_lines = [
        "📌 KEY INSIGHTS FROM MARKETING ANALYSIS (10 Questions)",
        "─────────────────────────────────────────────────────",
        "",
        "🔹 YOUR 5 CORE QUESTIONS:",
        "───────────────────────",
    ]

    # Q1: Top campaigns
    if 'Campaign Analysis' in data and 'Top Revenue' in data['Campaign Analysis']:
        tr = data['Campaign Analysis']['Top Revenue'].iloc[0]
        insights_lines.append(f"🏆 Top Revenue Campaign: {tr['Campaign_Name']} – Revenue ₹{tr['Revenue']:,.0f}, ROI {tr['ROI']:.1f}%")
    if 'Campaign Analysis' in data and 'Top Conversions' in data['Campaign Analysis']:
        tc = data['Campaign Analysis']['Top Conversions'].iloc[0]
        insights_lines.append(f"🎯 Top Conversions Campaign: {tc['Campaign_Name']} – Conversions {tc['Conversion']:,.0f}, CVR {tc['CVR']:.1f}%")

    # Q2: Channel performance
    if 'Campaign Analysis' in data and 'Channel Performance' in data['Campaign Analysis']:
        ch = data['Campaign Analysis']['Channel Performance']
        best_ch = ch.iloc[0]
        insights_lines.append(f"📢 Best Channel by Revenue: {best_ch['Channel']} – ₹{best_ch['Revenue']:,.0f}, ROI {best_ch['ROI']:.1f}%")
        # Add highest CTR, CVR, lowest CPA
        best_ctr = ch.loc[ch['CTR'].idxmax()]
        insights_lines.append(f"⚡ Highest CTR Channel: {best_ctr['Channel']} – CTR {best_ctr['CTR']:.1f}%")
        best_cvr = ch.loc[ch['CVR'].idxmax()]
        insights_lines.append(f"📈 Highest CVR Channel: {best_cvr['Channel']} – CVR {best_cvr['CVR']:.1f}%")
        best_cpa = ch.loc[ch['CPA'].idxmin()]
        insights_lines.append(f"💰 Lowest CPA Channel: {best_cpa['Channel']} – CPA ₹{best_cpa['CPA']:.0f}")

    # Q3: Regional impact
    if 'Campaign Analysis' in data and 'Regional Impact' in data['Campaign Analysis']:
        reg = data['Campaign Analysis']['Regional Impact']
        top_reg = reg.iloc[0]
        insights_lines.append(f"🌍 Top Region: {top_reg['Region']} – Revenue ₹{top_reg['Revenue']:,.0f}, ROI {top_reg['ROI']:.1f}%")

    # Q4: Audience responsiveness
    if 'Campaign Analysis' in data and 'Top Audiences' in data['Campaign Analysis']:
        aud = data['Campaign Analysis']['Top Audiences'].iloc[0]
        insights_lines.append(f"👥 Most Valuable Audience: {aud['Target_Audience']} – Revenue ₹{aud['Revenue']:,.0f}, ROI {aud['ROI']:.1f}%")

    # Q5: Spend vs revenue correlation
    if 'Campaign Analysis' in data and 'Spend vs Revenue' in data['Campaign Analysis']:
        corr_row = data['Campaign Analysis']['Spend vs Revenue']
        corr_val = corr_row.iloc[0]['Value']
        corr_text = corr_row.iloc[1]['Value']
        insights_lines.append(f"📊 Spend‑Revenue Correlation: {corr_val} – {corr_text}")

    insights_lines.append("")
    insights_lines.append("🔹 ADDITIONAL ML INSIGHTS (Our 5 Questions):")
    insights_lines.append("────────────────────────────────────────")

    # Q6: Campaign clustering
    if 'ML Insights' in data and 'Cluster Profiles' in data['ML Insights']:
        clust = data['ML Insights']['Cluster Profiles']
        high = clust[clust['Label'] == 'High ROI']
        if not high.empty:
            insights_lines.append(f"🗂️ Campaign Clusters: Found {len(clust)} clusters. High‑ROI cluster has {high['Size'].iloc[0]} campaigns (avg ROI {high['ROI'].iloc[0]:.1f}%).")
        else:
            insights_lines.append(f"🗂️ Campaign Clusters: Found {len(clust)} clusters. (No explicit High‑ROI label)")

    # Q7: Audience segmentation
    if 'Audience Segments' in data and 'Cluster Summary' in data['Audience Segments']:
        aud_clust = data['Audience Segments']['Cluster Summary']
        insights_lines.append(f"👥 Audience Segmentation: {len(aud_clust)} audience segments identified. High‑value segment comprises {aud_clust.loc[aud_clust['Label']=='High Value', 'Size'].values[0] if 'High Value' in aud_clust['Label'].values else '?'} audiences.")

    # Q8: Feature importance
    if 'ML Insights' in data and 'Feature Importance' in data['ML Insights']:
        fi = data['ML Insights']['Feature Importance'].head(3)
        top_features = ", ".join([f"{row['Feature']} ({row['Importance']:.2f})" for _, row in fi.iterrows()])
        insights_lines.append(f"🔑 Top 3 Features Driving Revenue: {top_features}")

    # Q9: Forecasting
    if 'Forecast' in data and 'Forecast' in data['Forecast']:
        fc = data['Forecast']['Forecast']
        total_next = fc['yhat'].sum()
        insights_lines.append(f"🔮 Forecast Next 90 Days: ₹{total_next:,.0f} total predicted revenue.")

    # Q10: Budget optimization
    if 'Budget Optimization' in data and 'Optimal Allocation' in data['Budget Optimization']:
        opt = data['Budget Optimization']['Optimal Allocation']
        top_alloc = opt.iloc[0]
        insights_lines.append(f"💰 Optimal Budget Allocation: Top channel {top_alloc['Channel']} gets ₹{top_alloc['Allocated Budget (₹)']:,.0f} (expected ROI {top_alloc['Expected ROI %']:.1f}%).")

    insights_lines.append("")
    insights_lines.append("💡 RECOMMENDATIONS:")
    insights_lines.append("• Focus budget on high‑ROI channels and campaigns.")
    insights_lines.append("• Re‑engage audiences from high‑value segments with tailored offers.")
    insights_lines.append("• Test creative types that perform best per channel.")
    insights_lines.append("• Use forecast to plan promotions ahead of peak months.")
    insights_lines.append("• Regularly retrain clustering models as new data arrives.")

    # Write insights to sheet
    for i, line in enumerate(insights_lines, start=1):
        cell = ws_insights.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14, color='1F4E79')
        elif line.startswith("🔹") or line.startswith("💡") or line.startswith("────────────────"):
            cell.font = Font(bold=True, size=12)
        else:
            cell.font = Font(size=11)
        cell.alignment = Alignment(horizontal='left')

    ws_insights.column_dimensions['A'].width = 100

    # ------------------------------------------------------------
    # 5. ADD ALL OTHER SHEETS FROM SOURCE FILES + CHARTS
    # ------------------------------------------------------------
    for source_name, sheets in data.items():
        for sheet_name, df in sheets.items():
            new_sheet_name = f"{source_name[:3]}_{sheet_name}"[:31]
            ws = wb.create_sheet(new_sheet_name)

            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.border = thin_border
                        if isinstance(value, (int, float)):
                            col_name = ws.cell(row=1, column=c_idx).value
                            if col_name in ['Revenue', 'Spend', 'CPA', 'Revenue_Per_Impression', 'Allocated Budget (₹)']:
                                cell.number_format = currency_format
                            elif col_name in ['ROI', 'CTR', 'CVR']:
                                cell.number_format = percent_format
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='left')

            # Auto‑fit columns
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        try:
                            max_len = max(max_len, len(str(cell.value)))
                        except:
                            pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

            # Add bar chart for key sheets
            if new_sheet_name in ['Cam_Channel Performance', 'Cam_Regional Impact', 'Cam_Top Revenue']:
                if 'Revenue' in df.columns:
                    val_col = list(df.columns).index('Revenue') + 1
                    cat_col = 1
                    chart_title = f"{sheet_name}"
                    add_chart(ws, chart_title, val_col, cat_col, 'E2')

    # ------------------------------------------------------------
    # 6. SAVE THE MASTER DASHBOARD
    # ------------------------------------------------------------
    wb.save(dashboard_path)
    print(f"\n✅ Professional master dashboard saved to:\n   {dashboard_path}")
    print("\n🎉 Your Marketing Masterpiece is now complete and includes insights from all 10 questions!")

except Exception as e:
    print("\n❌ AN ERROR OCCURRED:")
    traceback.print_exc()

finally:
    input("\nPress Enter to exit...")