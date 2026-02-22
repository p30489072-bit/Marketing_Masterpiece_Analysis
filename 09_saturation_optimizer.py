# scripts/09_saturation_optimizer.py
"""
===============================================================================
 LAYER 3: MARKETING MIX & SATURATION CURVES
===============================================================================
Adds to your Marketing Master Dashboard:
   • Channel Saturation – finds optimal spend per channel (diminishing returns)
   • Budget Optimizer – simple scenario planner
   • Scenario Planner – interactive (user enters budget, allocation updates)
===============================================================================
"""

import pandas as pd
import numpy as np
import os
import sqlite3
import traceback
from scipy.optimize import curve_fit
from sklearn.metrics import r2_score
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# HELPER FUNCTION – FORMAT INDIAN NUMBERS
# ------------------------------------------------------------
def format_indian(num):
    """Convert large numbers to Indian readable format (K, L, Cr)"""
    if pd.isna(num) or num == 0:
        return "0"
    if num >= 10**7:  # 1 Crore
        return f"{num/10**7:.1f}Cr"
    elif num >= 10**5:  # 1 Lakh
        return f"{num/10**5:.1f}L"
    elif num >= 10**3:  # 1 Thousand
        return f"{num/10**3:.1f}K"
    else:
        return str(int(num))

# ------------------------------------------------------------
# SATURATION MODEL (logarithmic)
# ------------------------------------------------------------
def log_func(x, a, b):
    return a * np.log(x) + b

try:
    print("=" * 60)
    print("📊 LAYER 3: CHANNEL SATURATION & BUDGET OPTIMIZER")
    print("=" * 60)

    # ------------------------------------------------------------
    # LOAD CLEANED DATA
    # ------------------------------------------------------------
    db_path = os.path.join('..', 'output', 'marketing.db')
    if os.path.exists(db_path):
        conn = sqlite3.connect(db_path)
        df = pd.read_sql_query("SELECT * FROM campaigns", conn)
        conn.close()
        print("✅ Loaded cleaned data from SQLite.")
    else:
        excel_path = os.path.join('..', 'output', 'cleaned_marketing_data.xlsx')
        df = pd.read_excel(excel_path)
        print("✅ Loaded cleaned data from Excel.")

    print(f"📊 Total rows: {len(df):,}")

    # ------------------------------------------------------------
    # PREPARE DATA PER CHANNEL
    # ------------------------------------------------------------
    channels = df['Channel'].unique()
    results = []

    for ch in channels:
        ch_data = df[df['Channel'] == ch].copy()
        ch_data = ch_data[(ch_data['Spend'] > 0) & (ch_data['Revenue'] > 0)]
        if len(ch_data) < 5:
            print(f"⚠️ Channel {ch} has only {len(ch_data)} data points – skipping curve fit.")
            continue

        ch_data['spend_bin'] = pd.qcut(ch_data['Spend'], q=10, duplicates='drop')
        binned = ch_data.groupby('spend_bin').agg({
            'Spend': 'mean',
            'Revenue': 'mean',
            'ROI': 'mean'
        }).reset_index()
        binned = binned.sort_values('Spend')

        x = binned['Spend'].values
        y = binned['Revenue'].values
        if len(x) < 3:
            continue

        try:
            popt, _ = curve_fit(log_func, x, y, maxfev=5000)
            a, b = popt
            y_pred = log_func(x, a, b)
            r2 = r2_score(y, y_pred)

            sim_spend = np.linspace(x.min() * 0.5, x.max() * 1.5, 100)
            sim_rev = log_func(sim_spend, a, b)
            sim_roi = (sim_rev - sim_spend) / sim_spend * 100

            optimal_idx = np.argmax(sim_roi)
            optimal_spend = sim_spend[optimal_idx]
            optimal_roi = sim_roi[optimal_idx]

            current_avg_spend = ch_data['Spend'].mean()
            current_avg_roi = ch_data['ROI'].mean()

            results.append({
                'Channel': ch,
                'Data_Points': len(ch_data),
                'R2': round(r2, 2),
                'Current_Avg_Spend': current_avg_spend,
                'Current_Avg_ROI': current_avg_roi,
                'Optimal_Spend': optimal_spend,
                'Optimal_ROI': optimal_roi,
                'Change_Spend': optimal_spend - current_avg_spend,
                'ROI_Lift': optimal_roi - current_avg_roi,
                'Recommendation': 'Increase' if optimal_spend > current_avg_spend else 'Decrease' if optimal_spend < current_avg_spend else 'Maintain'
            })
        except Exception as e:
            print(f"⚠️ Could not fit curve for {ch}: {e}")

    sat_df = pd.DataFrame(results)
    if sat_df.empty:
        raise ValueError("No saturation curves could be fitted. Try with more data per channel.")

    sat_df['Current_Avg_Spend_fmt'] = sat_df['Current_Avg_Spend'].apply(format_indian)
    sat_df['Optimal_Spend_fmt'] = sat_df['Optimal_Spend'].apply(format_indian)
    sat_df['Change_Spend_fmt'] = sat_df['Change_Spend'].apply(format_indian)
    sat_df['Current_Avg_ROI'] = sat_df['Current_Avg_ROI'].round(1)
    sat_df['Optimal_ROI'] = sat_df['Optimal_ROI'].round(1)
    sat_df['ROI_Lift'] = sat_df['ROI_Lift'].round(1)

    sat_df = sat_df.sort_values('ROI_Lift', ascending=False)

    # ------------------------------------------------------------
    # GENERATE INSIGHTS TEXT
    # ------------------------------------------------------------
    n_under = (sat_df['Recommendation'] == 'Increase').sum()
    n_over = (sat_df['Recommendation'] == 'Decrease').sum()
    n_optimal = (sat_df['Recommendation'] == 'Maintain').sum()
    best_opp = sat_df.iloc[0] if len(sat_df) > 0 else None
    avg_r2 = sat_df['R2'].mean()

    insights_lines = [
        "📊 SATURATION INSIGHTS",
        "─────────────────────────────────────",
        f"• {n_under} channel(s) are under‑spent – increase budget to capture higher ROI.",
        f"• {n_over} channel(s) are over‑spent – reduce to avoid diminishing returns.",
    ]
    if best_opp is not None:
        insights_lines.append(
            f"• Biggest opportunity: {best_opp['Channel']} – increase spend by {best_opp['Change_Spend_fmt']} "
            f"for +{best_opp['ROI_Lift']:.1f}% ROI lift."
        )
    if n_optimal > 0:
        optimal_channels = sat_df[sat_df['Recommendation'] == 'Maintain']['Channel'].tolist()
        insights_lines.append(f"• Already optimal: {', '.join(optimal_channels)} – maintain current spend.")
    insights_lines.append(f"• Average R² of fitted models: {avg_r2:.2f} – {'good reliability' if avg_r2>0.7 else 'moderate reliability'}.")
    insights_lines.append("")

    # ------------------------------------------------------------
    # LOAD EXISTING DASHBOARD AND ADD NEW SHEETS
    # ------------------------------------------------------------
    INPUT_PATH = os.path.join('..', 'output', 'Marketing_Master_Dashboard.xlsx')
    OUTPUT_PATH = INPUT_PATH

    if not os.path.exists(INPUT_PATH):
        raise FileNotFoundError(f"Dashboard file not found: {INPUT_PATH}. Run 07_dashboard.py first.")

    wb = load_workbook(INPUT_PATH)

    # Remove old sheets if they exist
    for sheet in ['Channel Saturation', 'Budget Optimizer', 'Scenario Planner']:
        if sheet in wb.sheetnames:
            wb.remove(wb[sheet])
            print(f"🗑️ Removed old sheet: {sheet}")

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    currency_format = '₹#,##0'
    percent_format = '0.00%'

    def format_sheet(ws):
        # Format headers (first row)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        # Auto‑fit columns – iterate by column index
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        # Borders for data cells (skip header row)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    # ================== SHEET 1: CHANNEL SATURATION ==================
    sat_display = sat_df[['Channel', 'Data_Points', 'R2', 'Current_Avg_Spend_fmt', 'Current_Avg_ROI',
                          'Optimal_Spend_fmt', 'Optimal_ROI', 'Change_Spend_fmt', 'ROI_Lift', 'Recommendation']].copy()
    sat_display.columns = ['Channel', 'Data Pts', 'R²', 'Current Spend', 'Current ROI %',
                           'Optimal Spend', 'Optimal ROI %', 'Change', 'ROI Lift %', 'Recommend']

    ws_sat = wb.create_sheet('Channel Saturation')

    # Write insights at the top
    for i, line in enumerate(insights_lines, start=1):
        cell = ws_sat.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14, color='1F4E79')
        elif line.startswith('─'):
            cell.font = Font(size=10)
        else:
            cell.font = Font(size=11)
        ws_sat.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

    # Write table
    start_row = len(insights_lines) + 2
    for r_idx, row in enumerate(dataframe_to_rows(sat_display, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_sat.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            else:
                if c_idx in [4,6,8]:
                    cell.number_format = currency_format
                elif c_idx in [5,7,9]:
                    cell.number_format = percent_format
                cell.border = thin_border

    if len(sat_df) > 0:
        roi_lift_col = 9
        data_bar = DataBarRule(start_type='min', end_type='max', color='63BE7B', showValue=True)
        ws_sat.conditional_formatting.add(
            f'{get_column_letter(roi_lift_col)}{start_row+1}:{get_column_letter(roi_lift_col)}{start_row+len(sat_df)}',
            data_bar
        )

    # ================== SHEET 2: BUDGET OPTIMIZER (static) ==================
    ws_opt = wb.create_sheet('Budget Optimizer')
    ws_opt['A1'] = '💰 BUDGET OPTIMIZER'
    ws_opt['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws_opt.merge_cells('A1:C1')

    ws_opt['A3'] = 'Channel'
    ws_opt['B3'] = 'Optimal Spend'
    ws_opt['C3'] = 'Expected ROI'
    for col in 'ABC':
        ws_opt[col+'3'].font = header_font
        ws_opt[col+'3'].fill = header_fill
        ws_opt[col+'3'].alignment = Alignment(horizontal='center')

    for i, row in sat_df.iterrows():
        r = 4 + i
        ws_opt[f'A{r}'] = row['Channel']
        ws_opt[f'B{r}'] = row['Optimal_Spend_fmt']
        ws_opt[f'C{r}'] = row['Optimal_ROI'] / 100
        ws_opt[f'C{r}'].number_format = percent_format

    format_sheet(ws_opt)

    # ================== SHEET 3: INTERACTIVE SCENARIO PLANNER ==================
    ws_scen = wb.create_sheet('Scenario Planner')

    ws_scen['A1'] = '🎯 INTERACTIVE BUDGET ALLOCATOR'
    ws_scen['A1'].font = Font(bold=True, size=14, color='1F4E79')
    ws_scen.merge_cells('A1:D1')

    ws_scen['A3'] = 'Enter Total Budget:'
    ws_scen['A3'].font = Font(bold=True)
    ws_scen['B3'] = 10000000  # default ₹1Cr
    ws_scen['B3'].number_format = currency_format

    ws_scen['A5'] = 'Channel'
    ws_scen['B5'] = 'Optimal Spend Ratio'
    ws_scen['C5'] = 'Allocated Budget'
    ws_scen['D5'] = 'Expected ROI'
    for col in 'ABCD':
        ws_scen[col+'5'].font = header_font
        ws_scen[col+'5'].fill = header_fill

    total_optimal = sat_df['Optimal_Spend'].sum()
    row = 6
    for i, ch in enumerate(sat_df['Channel']):
        ws_scen[f'A{row}'] = ch
        ratio = sat_df.iloc[i]['Optimal_Spend'] / total_optimal
        ws_scen[f'B{row}'] = ratio
        ws_scen[f'B{row}'].number_format = '0.00%'
        ws_scen[f'C{row}'] = f'=B$3*B{row}'
        ws_scen[f'C{row}'].number_format = currency_format
        ws_scen[f'D{row}'] = sat_df.iloc[i]['Optimal_ROI'] / 100
        ws_scen[f'D{row}'].number_format = percent_format
        row += 1

    # Total row
    ws_scen[f'A{row}'] = 'TOTAL'
    ws_scen[f'A{row}'].font = Font(bold=True)
    ws_scen[f'C{row}'] = f'=SUM(C6:C{row-1})'
    ws_scen[f'C{row}'].number_format = currency_format
    ws_scen[f'C{row}'].font = Font(bold=True)

    # Format columns
    ws_scen.column_dimensions['A'].width = 20
    ws_scen.column_dimensions['B'].width = 20
    ws_scen.column_dimensions['C'].width = 20
    ws_scen.column_dimensions['D'].width = 20

    # Apply borders to the table area
    for r in range(5, row+1):
        for c in range(1,5):
            cell = ws_scen.cell(row=r, column=c)
            cell.border = thin_border

    # -------------------- SAVE --------------------
    wb.save(OUTPUT_PATH)
    print(f"✅ Saturation, Optimizer, and Scenario Planner sheets added to:\n   {OUTPUT_PATH}")

    # Console insights
    print("\n" + "=" * 60)
    print("🔑 SATURATION INSIGHTS")
    print("=" * 60)
    for _, row in sat_df.head(3).iterrows():
        print(f"\n📢 {row['Channel']}:")
        print(f"   Current spend: {format_indian(row['Current_Avg_Spend'])} → ROI {row['Current_Avg_ROI']:.1f}%")
        print(f"   Optimal spend: {format_indian(row['Optimal_Spend'])} → ROI {row['Optimal_ROI']:.1f}%")
        print(f"   Change: {format_indian(row['Change_Spend'])} (ROI lift {row['ROI_Lift']:.1f}%)")
        print(f"   Recommendation: {row['Recommendation']} spend")

    print("\n✅ Layer 3 complete! Open the dashboard and try the 'Scenario Planner' sheet – change the budget in B3 and watch allocations update!")

except Exception as e:
    print("\n❌ AN ERROR OCCURRED:")
    traceback.print_exc()

finally:
    input("\nPress Enter to close this window...")