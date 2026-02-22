# scripts/11_anomaly_detection.py
"""
===============================================================================
 LAYER 5: ANOMALY DETECTION – FLAG UNUSUAL CAMPAIGNS
===============================================================================
Adds to your Marketing Master Dashboard:
   • Anomaly Detection – identifies campaigns with abnormal metrics
   • Helps catch fraud, errors, or unexpected performance
===============================================================================
"""

import pandas as pd
import numpy as np
import os
import sqlite3
import traceback
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ------------------------------------------------------------
# HELPER FUNCTION – FORMAT INDIAN NUMBERS
# ------------------------------------------------------------
def format_indian(num):
    if pd.isna(num) or num == 0:
        return "0"
    if num >= 10**7:
        return f"{num/10**7:.1f}Cr"
    elif num >= 10**5:
        return f"{num/10**5:.1f}L"
    elif num >= 10**3:
        return f"{num/10**3:.1f}K"
    else:
        return str(int(num))

try:
    print("=" * 60)
    print("📊 LAYER 5: ANOMALY DETECTION")
    print("=" * 60)

    # ------------------------------------------------------------
    # LOAD CLEANED DATA
    # ------------------------------------------------------------
    db_path = os.path.join('..', 'output', 'marketing.db')
    if os.path.exists(db_path):
        conn = sqlite3.connect(db_path)
        df = pd.read_sql_query("SELECT * FROM campaigns", conn)
        conn.close()
        print("✅ Loaded cleaned data from SQLite.")
    else:
        excel_path = os.path.join('..', 'output', 'cleaned_marketing_data.xlsx')
        df = pd.read_excel(excel_path)
        print("✅ Loaded cleaned data from Excel.")

    print(f"📊 Total rows: {len(df):,}")

    # ------------------------------------------------------------
    # PREPARE FEATURES FOR ANOMALY DETECTION
    # ------------------------------------------------------------
    # Use numeric columns that could indicate anomalies
    feature_cols = ['ROI', 'CTR', 'CVR', 'CPA', 'Spend', 'Revenue', 'Clicks', 'Impressions']
    # Ensure all exist
    feature_cols = [col for col in feature_cols if col in df.columns]

    X = df[feature_cols].copy()
    # Replace inf/-inf with NaN, then drop
    X = X.replace([np.inf, -np.inf], np.nan)
    X = X.dropna()

    if len(X) < len(df):
        print(f"⚠️ Dropped {len(df)-len(X)} rows with missing/infinite values.")

    # Standardize features
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)

    # ------------------------------------------------------------
    # ISOLATION FOREST
    # ------------------------------------------------------------
    iso_forest = IsolationForest(contamination=0.05, random_state=42)  # assume ~5% anomalies
    preds = iso_forest.fit_predict(X_scaled)
    scores = iso_forest.decision_function(X_scaled)

    # -1 = anomaly, 1 = normal
    df_anomaly = df.loc[X.index].copy()
    df_anomaly['Anomaly'] = preds
    df_anomaly['Anomaly_Score'] = scores
    df_anomaly['Anomaly_Label'] = df_anomaly['Anomaly'].map({1: 'Normal', -1: '⚠️ Anomaly'})

    # Identify which metrics contributed most (for insight)
    # We can look at how far from mean each metric is
    feature_means = X.mean()
    feature_stds = X.std()
    normalized_diff = (X - feature_means) / feature_stds
    # For anomalies, take absolute deviation
    anomaly_indices = df_anomaly[df_anomaly['Anomaly'] == -1].index
    if len(anomaly_indices) > 0:
        top_contributors = normalized_diff.loc[anomaly_indices].abs().mean().sort_values(ascending=False).head(3)
        top_features = top_contributors.index.tolist()
    else:
        top_features = []

    # ------------------------------------------------------------
    # SUMMARY STATS
    # ------------------------------------------------------------
    total_anomalies = (df_anomaly['Anomaly'] == -1).sum()
    pct_anomalies = total_anomalies / len(df_anomaly) * 100

    # ------------------------------------------------------------
    # GENERATE INSIGHTS
    # ------------------------------------------------------------
    insights_lines = [
        "🚨 ANOMALY INSIGHTS",
        "─────────────────────────────────────",
        f"• Detected {total_anomalies} anomalous campaigns ({pct_anomalies:.1f}% of total).",
    ]
    if top_features:
        insights_lines.append(f"• Main drivers: {', '.join(top_features)} (extreme values).")
    insights_lines.append("• Review flagged campaigns – they may indicate errors, fraud, or exceptional performance.")
    insights_lines.append("")

    # ------------------------------------------------------------
    # LOAD DASHBOARD AND ADD NEW SHEET
    # ------------------------------------------------------------
    INPUT_PATH = os.path.join('..', 'output', 'Marketing_Master_Dashboard.xlsx')
    OUTPUT_PATH = INPUT_PATH

    if not os.path.exists(INPUT_PATH):
        raise FileNotFoundError(f"Dashboard file not found: {INPUT_PATH}. Run 07_dashboard.py first.")

    wb = load_workbook(INPUT_PATH)

    # Remove old anomaly sheet if exists
    if 'Anomaly Detection' in wb.sheetnames:
        wb.remove(wb['Anomaly Detection'])
        print("🗑️ Removed old Anomaly Detection sheet.")

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    currency_format = '₹#,##0'
    percent_format = '0.00%'

    def format_sheet(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    # Create new sheet
    ws_anom = wb.create_sheet('Anomaly Detection')

    # Write insights at top
    for i, line in enumerate(insights_lines, start=1):
        cell = ws_anom.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14, color='1F4E79')
        elif line.startswith('─'):
            cell.font = Font(size=10)
        else:
            cell.font = Font(size=11)
        ws_anom.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    # Write anomaly table
    start_row = len(insights_lines) + 2
    cols_display = ['Campaign_Id', 'Campaign_Name', 'Channel', 'ROI', 'CTR', 'CVR', 'Spend', 'Revenue', 'Anomaly_Label']
    # Only show columns that exist
    cols_display = [c for c in cols_display if c in df_anomaly.columns]
    disp_df = df_anomaly[cols_display].copy()
    # Format ROI, CTR, CVR as percentages
    for col in ['ROI', 'CTR', 'CVR']:
        if col in disp_df.columns:
            disp_df[col] = disp_df[col].round(1)
    # Format Spend, Revenue
    for col in ['Spend', 'Revenue']:
        if col in disp_df.columns:
            disp_df[col] = disp_df[col].apply(format_indian)

    # Write headers
    for c_idx, col_name in enumerate(disp_df.columns, start=1):
        cell = ws_anom.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Write data rows
    for r_idx, (_, row) in enumerate(disp_df.iterrows(), start=start_row+1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_anom.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            # Conditional formatting for anomaly label
            col_name = disp_df.columns[c_idx-1]
            if col_name == 'Anomaly_Label' and '⚠️' in str(value):
                cell.font = Font(bold=True, color='9C0006')
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            # Number formats
            if col_name in ['ROI', 'CTR', 'CVR']:
                cell.number_format = percent_format

    # Add conditional formatting for numeric columns to highlight extremes
    # (optional, but we can leave for now)

    format_sheet(ws_anom)

    wb.save(OUTPUT_PATH)
    print(f"✅ Anomaly Detection sheet added to:\n   {OUTPUT_PATH}")

except Exception as e:
    print("\n❌ AN ERROR OCCURRED:")
    traceback.print_exc()

finally:
    input("\nPress Enter to close this window...")